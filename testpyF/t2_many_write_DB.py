'''
格式为
日期 开盘价 收盘价 当日最高 当日最低 成交量 成交额 振幅 涨跌幅 涨跌额 换手

'''
import time
import random

import akshare as ak
import openpyxl
import pymysql
from concurrent.futures import ThreadPoolExecutor

stock_num1 = '600584'
def read_stock_num():
    wb = openpyxl.load_workbook('stock_num.xlsx')
    sheet = wb[list(wb.sheetnames)[0]]
    NumList = []
    for i in range(1, sheet.max_row + 1):
        NumList.append(sheet['a' + str(i)].value)

    return NumList


def main():
    NumList = read_stock_num()
    for stock_num in NumList:
        create_tabel(stock_num)
        try:
            datas = extract_data(stock_num)
            insert_table(stock_num, datas)
        except:
            print(f'{stock_num}上市时间过短')
            continue
        time.sleep(random.random() * 2)


def create_tabel(stock_num):
    conn = pymysql.connect(host='localhost', port=3306, user='zengke', password='zk.123456', database='stock_data',
                           charset='utf8mb4')
    try:
        with conn.cursor() as cursor:  # type: Cursor

            sql = f'''create table if not exists tb_{stock_num} (st_id int not null auto_increment comment '标号',
                    dates date comment '日期',
                    open_val decimal(4,2) comment '开盘价',
                    close_val decimal(4,2) comment '收盘价',
                    highest decimal(4,2) comment '最高价',
                    lowest decimal(4,2) comment '最低价',
                    volume decimal(10,0) comment '成交量',
                    volume_val decimal(14,0) comment '成交额',
                    amplitude decimal(4,2) comment '振幅',
                    涨跌幅 decimal(4,2) comment '涨跌幅',
                    涨跌额 decimal(4,2) comment '涨跌额',
                    换手 decimal(4,2) comment '换手率',				
                    primary key (st_id)
                    
                    
                    )engine=innodb  comment '股票表';
                    
                    '''

            cursor.execute(sql)
            conn.commit()
            print(f'表tb_{stock_num}读取成功')
    except pymysql.MySQLError as err:
        print(err)
    finally:
        print('创建结束')
        conn.close()


def insert_table(stock_num, data):
    conn = pymysql.connect(host='localhost', port=3306, user='zengke', password='zk.123456', database='stock_data',
                           charset='utf8mb4')
    try:
        with conn.cursor() as cursor:  # type: Cursor

            cursor.executemany(f'''insert into tb_{stock_num} (dates,open_val,close_val,highest,lowest,volume,volume_val,amplitude,涨跌幅,涨跌额,换手)
                        values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)

                    ''', data)

            conn.commit()
    except pymysql.MySQLError as err:
        print(err)
    finally:
        print('加入结束')
        conn.close()


def extract_data(stock_num):
    data_list = []
    stock_zh_a_hist_df = ak.stock_zh_a_hist(symbol=stock_num, period="daily", start_date="20180301",
                                            end_date='20221118',
                                            adjust="qfq")  # symbol:股票代码 ,period:时间段(天,周,月) ,start_date:开始时间.end_date:结束时间,adjust:复权
    for item in stock_zh_a_hist_df.values:
        data_list.append(tuple(item))

    return data_list


if __name__ == '__main__':
    start = time.time()

    main()
    end = time.time()
    print(f'总共花费{end - start}秒')
