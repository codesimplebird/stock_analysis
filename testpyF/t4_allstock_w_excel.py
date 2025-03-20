import akshare as ak
import openpyxl

from concurrent.futures import ThreadPoolExecutor


def main():
    codes = read_stock_num()
    with ThreadPoolExecutor() as pool:
        datas = pool.map(extract_num, codes)

    exc_add(datas)


def read_stock_num():
    wb = openpyxl.load_workbook('stock_num.xlsx')
    sheet = wb[list(wb.sheetnames)[0]]
    NumList = []
    for i in range(1, sheet.max_row + 1):
        NumList.append(sheet['a' + str(i)].value)

    return NumList


def extract_num(code):
    try:
        datas = ak.stock_individual_info_em(symbol=f"{code}")
        data = datas.values[:, 1]
        print(code)
    except:
        print("出错啦")
        return list(f'{code}出错')
    return data


def exc_add(datalist):
    wb = openpyxl.load_workbook('stock_num.xlsx')
    name = "one_stock_info"
    wb.create_sheet(title=name)
    sheet = wb[name]
    sheet.append(('总市值', "流通市值", '行业', '上市日期', '代码', '名称', '总股本', '流通股本'))
    for item in datalist:
        sheet.append(tuple(item))
    wb.save('stock_num.xlsx')


if __name__ == '__main__':
    main()
