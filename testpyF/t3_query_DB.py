import time

import pymysql
import openpyxl

'''query 询问'''
'''获取数据库数据'''

def main():
    stockList = read_stock_num()
    datalist = select_data(stockList)

    write_xl(datalist)


def read_stock_num():
    wb = openpyxl.load_workbook(r'stock_num.xlsx')
    sheet = wb[list(wb.sheetnames)[0]]
    NumList = []
    for i in range(1, sheet.max_row + 1):
        NumList.append(sheet['a' + str(i)].value)

    return NumList


def select_data(stock_num_list):
    conn = pymysql.connect(host='localhost', port=3306, user='zengke', password='zk.123456', database='stock_data',
                           charset='utf8mb4')
    datalist = []
    for stock_num in stock_num_list:

        try:
            with conn.cursor() as cur:  # type Cursor
                data = []
                data.append(stock_num)
                sql = f'''select distinct dates,  highest
                    from tb_{stock_num}
                    where dates between '2021-02-20' and '2022-11-17' and highest=
                    (select max(highest) from tb_{stock_num} 
                    where dates between '2021-02-20' and '2022-11-17')
                    union
                    select distinct dates,  highest
                    from tb_{stock_num}
                    where dates betw
                    een '2021-02-20' and '2022-11-17' and highest=
                    (select min(highest) from tb_{stock_num} 
                    where dates between '2021-02-20' and '2022-11-17');
                    '''
                # sql=f'''select dates,highest
                #         from tb_{stock_num}
                #         where dates between '2021-02-20' and '2022-11-17'
                #
                # '''
                cur.execute(sql)  # 执行语句
                while row := cur.fetchone():
                    data.append(row[0])
                    data.append(row[1])
                conn.commit()  # 提交
                datalist.append(data)

        except pymysql.MySQLError as err:
            print(err)

            conn.rollback()
            continue
    return datalist


def write_xl(datalist):
    wb = openpyxl.load_workbook('Share_data_xl/stockQuantify.xlsx')
    sheet = wb[list(wb.sheetnames)[0]]
    sheet.append(('代码', '最高价日期', '最高价', '最低价日期', '最低价'))
    for item in datalist:
        sheet.append(tuple(item))
    wb.save('Share_data_xl/stockQuantify.xlsx')


if __name__ == '__main__':
    start=time.time()
    main()
    end=time.time()
    print(f'使用了{end-start}')