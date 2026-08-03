import os
from datetime import datetime
import pymysql
import openpyxl


def main():
    stocklist = read_stock_num()
    datas = read_DB(stocklist)
    write_ex(datas)


def read_stock_num():
    wb = openpyxl.load_workbook(r'stockQuantify.xlsx')
    sheet = wb[list(wb.sheetnames)[0]]
    NumList = []
    for i in range(2, sheet.max_row + 1):
        data = []
        if sheet['a' + str(i)].value != None:
            data.append(sheet['a' + str(i)].value)
            data.append(sheet['b' + str(i)].value)
            data.append(sheet['d' + str(i)].value)
            NumList.append(data)
        else:
            break

    print("璇诲彇瀹屾垚")




    return NumList


def read_DB(stock_num_list):
    conn = pymysql.connect(host='localhost', port=3306, user=os.environ.get('MYSQL_USER', ''), password=os.environ.get('MYSQL_PASSWORD', ''), database='stock_data',
                           charset='utf8mb4')
    datalist = []
    for stock_num in stock_num_list:
        num=stock_num[0]
        date1=stock_num[1].strftime('%Y-%m-%d')
        date2=stock_num[2].strftime('%Y-%m-%d')

        try:
            with conn.cursor() as cur:  # type Cursor
                data=[]
                data.append(stock_num[0])
                sql = f'''  select avg(鎹㈡墜)
                            from tb_{num}
                            where st_id between (
                            select st_id
                            from tb_{num}
                            where dates='{date1}')-5 and
                            (select st_id
                            from tb_{num}
                            where dates='{date1}' )+5
                            union
                            select avg(鎹㈡墜)
                            from tb_{num}
                            where st_id between (
                            select st_id
                            from tb_{num}
                            where dates='{date2}')-5 and
                            (select st_id
                            from tb_{num}
                            where dates='{date2}' )+5;
                                    '''
                # sql=f'''select dates,highest
                #         from tb_{stock_num}
                #         where dates between '2021-02-20' and '2022-11-17'
                #
                # '''
                cur.execute(sql)  # 鎵ц璇彞
                while row := cur.fetchone():
                    data.append(row[0])

                print(data)

                print(f'鑾峰彇{num}瀹屾垚')
                conn.commit()  # 鎻愪氦
                datalist.append(data)

        except pymysql.MySQLError as err:
            print(err)

            conn.rollback()
            continue
    return datalist


def write_ex(dataslist):
    print('寮€濮嬪啓鍏ユ枃浠?)
    wb = openpyxl.load_workbook(r'stockQuantify.xlsx')
    name = '椤剁偣鍜屾渶浣庣偣鐨勯噺鑳藉叧绯?2'
    wb.create_sheet(title=name)
    sheet = wb[name]
    for item in dataslist:
        sheet.append(tuple(item))
    wb.save(r'stockQuantify.xlsx')


if __name__ == '__main__':
    # read_stock_num()
    main()